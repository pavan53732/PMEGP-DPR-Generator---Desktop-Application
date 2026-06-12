import json
import sys
try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed")
    sys.exit(1)

wb = load_workbook(r"audit-output/DPRPACKAGE.xlsx", data_only=False)

output = {}

# 1. Sheet level metadata (Print Area, Hidden Rows/Cols)
output["sheet_metadata"] = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    hidden_rows = [r for r, dim in ws.row_dimensions.items() if dim.hidden]
    hidden_cols = [c for c, dim in ws.column_dimensions.items() if dim.hidden]
    output["sheet_metadata"][sheet_name] = {
        "print_area": ws.print_area,
        "hidden_rows": hidden_rows,
        "hidden_cols": hidden_cols
    }

# 2. Data Validation Rules (Drop-downs)
output["data_validations"] = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    validations = []
    for dv in ws.data_validations.dataValidation:
        validations.append({
            "sqref": str(dv.sqref),
            "type": dv.type,
            "formula1": dv.formula1,
            "formula2": dv.formula2,
            "allow_blank": dv.allow_blank
        })
    if validations:
        output["data_validations"][sheet_name] = validations

# 3. Defined Names (Named Ranges)
output["defined_names"] = {}
for name, dn in wb.defined_names.items():
    output["defined_names"][name] = dn.attr_text

# 4. Unlocked Cells in DataSheet (True Input Fields)
ws_data = wb["DataSheet"]
unlocked_cells = []
for row in ws_data.iter_rows():
    for cell in row:
        if not cell.protection.locked:
            unlocked_cells.append({
                "cell": cell.coordinate,
                "value": str(cell.value),
                "is_formula": str(cell.value).startswith("=") if cell.value else False,
                "data_type": cell.data_type
            })
output["unlocked_cells_datasheet"] = unlocked_cells

# 5. All hardcoded non-empty, non-string cells in DataSheet (potential inputs missed by lock check)
potential_inputs = []
for row in ws_data.iter_rows():
    for cell in row:
        if cell.value is not None and not str(cell.value).startswith("=") and cell.data_type in ['n', 'd']:
            potential_inputs.append({
                "cell": cell.coordinate,
                "value": cell.value,
                "locked": cell.protection.locked
            })
output["numeric_hardcoded_datasheet"] = potential_inputs

with open("DPRPACKAGE-deeper-audit.json", "w", encoding="utf-8") as f:
    json.dump(output, f, indent=2)

print("Deeper audit complete. Output saved to DPRPACKAGE-deeper-audit.json")
