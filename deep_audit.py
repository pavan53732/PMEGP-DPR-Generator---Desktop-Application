"""Deep audit script to extract specific cell context from the XLSX workbook."""
import json
import sys
try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed")
    sys.exit(1)

wb = load_workbook(r"audit-output/DPRPACKAGE.xlsx", data_only=False)

def cell_info(sheet_name, cell_ref):
    """Get cell value, formula, and type."""
    ws = wb[sheet_name]
    cell = ws[cell_ref]
    return {
        "cell": f"{sheet_name}!{cell_ref}",
        "value": str(cell.value) if cell.value is not None else "EMPTY",
        "data_type": cell.data_type,
        "is_formula": str(cell.value).startswith("=") if cell.value else False,
    }

def row_context(sheet_name, row_num, col_start=1, col_end=22):
    """Get all non-empty cells in a row."""
    ws = wb[sheet_name]
    result = {}
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row_num, column=col)
        if cell.value is not None:
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col)
            result[f"{col_letter}{row_num}"] = {
                "value": str(cell.value),
                "type": cell.data_type,
            }
    return result

def area_context(sheet_name, row_start, row_end, col_start=1, col_end=22):
    """Get all non-empty cells in an area."""
    results = {}
    for row in range(row_start, row_end + 1):
        ctx = row_context(sheet_name, row, col_start, col_end)
        if ctx:
            results[f"row_{row}"] = ctx
    return results

print("=" * 80)
print("1. DPR_FRONT sheet - full content")
print("=" * 80)
front = area_context("DPR_FRONT", 1, 44, 1, 9)
print(json.dumps(front, indent=2))

print("\n" + "=" * 80)
print("2. Project_Report rows around broken cells (G14, J20, H21, H22)")
print("=" * 80)
pr_context = area_context("Project_Report", 1, 30, 1, 14)
print(json.dumps(pr_context, indent=2))

print("\n" + "=" * 80)
print("3. DataSheet L/M column area (rows 50-90) - lookup tables")
print("=" * 80)
ds_lookup = area_context("DataSheet", 50, 90, 12, 18)
print(json.dumps(ds_lookup, indent=2))

print("\n" + "=" * 80)
print("4. DPR_print rows 90-100 (around B94 #REF)")
print("=" * 80)
dpr_b94 = area_context("DPR_print", 85, 100, 1, 10)
print(json.dumps(dpr_b94, indent=2))

print("\n" + "=" * 80)
print("5. DPR_print rows 320-400 (broken formula area)")
print("=" * 80)
dpr_broken = area_context("DPR_print", 320, 400, 1, 10)
print(json.dumps(dpr_broken, indent=2))

print("\n" + "=" * 80)
print("6. DataSheet rows 30-40 (around M36)")
print("=" * 80)
ds_m36 = area_context("DataSheet", 30, 40, 1, 18)
print(json.dumps(ds_m36, indent=2))

print("\n" + "=" * 80)
print("7. DataSheet rows 1-35 (input fields)")
print("=" * 80)
ds_input = area_context("DataSheet", 1, 35, 1, 18)
print(json.dumps(ds_input, indent=2))

print("\n" + "=" * 80)
print("8. Merged ranges for DPR_FRONT")
print("=" * 80)
ws_front = wb["DPR_FRONT"]
for mr in ws_front.merged_cells.ranges:
    print(f"  {mr}")

print("\n" + "=" * 80)
print("9. Merged ranges for Project_Report rows 1-30")
print("=" * 80)
ws_pr = wb["Project_Report"]
for mr in ws_pr.merged_cells.ranges:
    bounds = mr.bounds  # (min_col, min_row, max_col, max_row)
    if bounds[1] <= 30:
        print(f"  {mr}")

print("\n" + "=" * 80)
print("10. DPR_print rows 120-135 (means of financing)")
print("=" * 80)
dpr_fin = area_context("DPR_print", 120, 135, 1, 10)
print(json.dumps(dpr_fin, indent=2))

print("\n" + "=" * 80)
print("11. DataSheet column M values (M55, M59, M64, M67, M70, M80, M83, M91)")
print("=" * 80)
target_cells = ["M55", "M56", "M59", "M64", "M67", "M70", "M80", "M83", "M91"]
for c in target_cells:
    info = cell_info("DataSheet", c)
    print(f"  {c}: {info}")

print("\n" + "=" * 80)
print("12. DataSheet rows 83-93 (financing + lookup)")
print("=" * 80)
ds_fin = area_context("DataSheet", 83, 93, 1, 18)
print(json.dumps(ds_fin, indent=2))

print("\nDONE")
