import json
import re
from openpyxl import load_workbook
from openpyxl.formula.tokenizer import Tokenizer

print("Loading workbook...")
wb = load_workbook(r"audit-output/DPRPACKAGE.xlsx", data_only=False)

formula_map = {}
dependencies = {}

# Regex to find standard A1 or SheetName!A1 references
ref_pattern = re.compile(r"('?[a-zA-Z0-9_]+'?!)?\$?[A-Z]+\$?[0-9]+")

print("Extracting formulas...")
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    formula_map[sheet_name] = {}
    for row in ws.iter_rows():
        for cell in row:
            val = str(cell.value) if cell.value is not None else ""
            if val.startswith("="):
                coord = cell.coordinate
                # Basic tokenization
                tokens = []
                try:
                    tok = Tokenizer(val)
                    for t in tok.items:
                        if t.type == "OPERAND" and t.subtype == "RANGE":
                            tokens.append(t.value)
                except:
                    # Fallback regex if tokenizer fails
                    tokens = [m.group(0) for m in ref_pattern.finditer(val)]
                
                formula_map[sheet_name][coord] = {
                    "formula": val,
                    "depends_on": list(set(tokens))
                }

# Now let's trace exactly what DataSheet inputs drive what outputs.
print("Mapping DataSheet dependencies...")
datasheet_links = {}

for sheet, cells in formula_map.items():
    if sheet == "DataSheet": continue
    for coord, data in cells.items():
        deps = data["depends_on"]
        for d in deps:
            if "DataSheet!" in d or "Datasheet!" in d or "datasheet!" in d:
                # normalize to base cell reference
                clean_ref = d.split("!")[-1].replace("$", "")
                if clean_ref not in datasheet_links:
                    datasheet_links[clean_ref] = []
                datasheet_links[clean_ref].append(f"{sheet}!{coord}")

# Look at logic specifically inside DataSheet (e.g. M column codes mapping to subsidies)
datasheet_internal_logic = {}
for coord, data in formula_map["DataSheet"].items():
    if coord.startswith("L") or coord.startswith("M") or coord.startswith("Q") or coord.startswith("R"):
        datasheet_internal_logic[coord] = data["formula"]

# Output results
output = {
    "DataSheet_Internal_Logic_Cells": datasheet_internal_logic,
    "Output_Sheets_Depending_On_DataSheet": datasheet_links,
    "Sample_DPR_Print_Formulas": {k: formula_map["DPR_print"][k] for k in list(formula_map["DPR_print"].keys())[:50]},
    "Project_Report_Formulas": {k: formula_map["Project_Report"][k] for k in list(formula_map["Project_Report"].keys())[:50]}
}

print("Saving logic-graph.json...")
with open("logic-graph.json", "w", encoding="utf-8") as f:
    json.dump(output, f, indent=2)

print("Done!")
